#!/usr/bin/env python3
"""
Export LA Adventure Club signups to a local Excel (.xlsx) file.

USAGE
    python3 tools/export-signups.py                      # Juneteenth walk (default)
    python3 tools/export-signups.py all                  # every event, one sheet
    python3 tools/export-signups.py solstice-canyon-hike # a specific event slug

CREDENTIALS  (read from tools/.env  — copy tools/.env.example)
    SUPABASE_URL          defaults to the club's project
    SUPABASE_SERVICE_KEY  REQUIRED — your service_role / secret key

The service key bypasses Row-Level Security so it can read EVERY signup.
KEEP IT SECRET: it lives only in tools/.env (git-ignored) and is never committed.
The spreadsheet it produces contains members' personal info — keep it private.
By default the file is written to the parent folder (outside the public repo);
set OUT_DIR to choose another location, e.g.  OUT_DIR=~/Desktop
"""

import os, sys, json, urllib.request, urllib.parse, datetime, subprocess

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXP = {'first_timer': 'First-timer', 'some': 'Some experience', 'experienced': 'Experienced'}


def load_env():
    envp = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    if os.path.exists(envp):
        for raw in open(envp, encoding='utf-8'):
            s = raw.strip()
            if not s or s.startswith('#') or '=' not in s:
                continue
            k, v = s.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def api_get(base, key, path, params):
    url = base.rstrip('/') + '/rest/v1/' + path + '?' + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={
        'apikey': key, 'Authorization': 'Bearer ' + key, 'Accept': 'application/json'})
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.load(r)
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', 'ignore')
        sys.exit(f"\nAPI error {e.code} on {path}:\n  {body}\n"
                 "(If this says 'permission denied', your key is the publishable one, "
                 "not the service_role secret key.)")


def fmt_ts(s):
    if not s:
        return ''
    try:
        return datetime.datetime.fromisoformat(s.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
    except Exception:
        return s


def main():
    load_env()
    base = os.environ.get('SUPABASE_URL', 'https://jmeswonsawyphrutjvuv.supabase.co')
    key = os.environ.get('SUPABASE_SERVICE_KEY')
    if not key:
        sys.exit("ERROR: SUPABASE_SERVICE_KEY is not set.\n"
                 "→ Copy tools/.env.example to tools/.env and paste your service_role key.\n"
                 "  (Supabase → Project Settings → API → service_role / secret key)")

    slug = sys.argv[1] if len(sys.argv) > 1 else 'juneteenth-point-vicente'

    # 1. events
    ev_params = {'select': 'id,slug,title,event_date,location,start_time'}
    if slug != 'all':
        ev_params['slug'] = 'eq.' + slug
    events = api_get(base, key, 'events', ev_params)
    if not events:
        sys.exit(f"No event found for slug '{slug}'. Try: python3 tools/export-signups.py all")
    ev_by_id = {e['id']: e for e in events}

    # 2. registrations for those events
    regs = api_get(base, key, 'registrations', {
        'select': 'user_id,event_id,guests,status,created_at',
        'order': 'created_at.asc',
        'event_id': 'in.(%s)' % ','.join(ev_by_id.keys())})
    regs = [r for r in regs if r.get('status') != 'cancelled']

    # 3. profiles for the registrants
    uids = sorted({r['user_id'] for r in regs})
    profiles = {}
    if uids:
        rows = api_get(base, key, 'profiles', {
            'select': 'id,full_name,email,phone,age_range,gender,'
                      'emergency_contact_name,emergency_contact_phone,experience_level',
            'id': 'in.(%s)' % ','.join(uids)})
        profiles = {p['id']: p for p in rows}

    write_output(slug, events, regs, ev_by_id, profiles)


def build_rows(regs, ev_by_id, profiles):
    headers = ['#', 'Name', 'Email', 'Phone', 'Age range', 'Gender',
               'Emergency contact', 'Emergency phone', 'Experience',
               'Guests', 'Party size', 'Status', 'Event', 'Event date', 'Signed up']
    out = []
    for i, r in enumerate(regs, 1):
        p = profiles.get(r['user_id'], {})
        e = ev_by_id.get(r['event_id'], {})
        g = r.get('guests') or 0
        out.append([
            i, p.get('full_name', ''), p.get('email', ''), p.get('phone', ''),
            p.get('age_range', ''), p.get('gender', ''),
            p.get('emergency_contact_name', ''), p.get('emergency_contact_phone', ''),
            EXP.get(p.get('experience_level', ''), p.get('experience_level', '') or ''),
            g, 1 + g, r.get('status', ''),
            e.get('title', ''), e.get('event_date', ''), fmt_ts(r.get('created_at'))])
    return headers, out


def write_output(slug, events, regs, ev_by_id, profiles):
    headers, rows = build_rows(regs, ev_by_id, profiles)
    title = events[0]['title'] if (slug != 'all' and events) else 'All events'
    out_dir = os.path.expanduser(os.environ.get('OUT_DIR') or os.path.dirname(REPO_ROOT))
    stem = ('all' if slug == 'all' else slug) + '-signups'

    # Prefer a real .xlsx; fall back to .csv if openpyxl can't be installed.
    try:
        try:
            import openpyxl  # noqa
        except ImportError:
            print('Installing openpyxl (one-time)…')
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', '--quiet', 'openpyxl'])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook(); ws = wb.active
        ws.title = (slug[:28] if slug != 'all' else 'All events')
        ws.append([f'LA Adventure Club — Signups: {title}'])
        ws.append([f"Generated {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} · "
                   f"{len(rows)} signup(s)"])
        ws.append([])
        ws.append(headers)
        for row in rows:
            ws.append(row)

        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'].font = Font(italic=True, color='888888')
        fill = PatternFill('solid', fgColor='E2531D')
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=c)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = fill
            cell.alignment = Alignment(vertical='center')
        ws.freeze_panes = 'A5'
        for i, w in enumerate([4, 20, 26, 15, 11, 12, 22, 16, 15, 8, 10, 11, 28, 12, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        out = os.path.join(out_dir, stem + '.xlsx')
        wb.save(out)
    except Exception as ex:
        import csv
        out = os.path.join(out_dir, stem + '.csv')
        with open(out, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)
        print(f'(openpyxl unavailable — wrote CSV instead: {ex})')

    print(f"\n✓ {len(rows)} signup(s) → {out}\n")


if __name__ == '__main__':
    main()
